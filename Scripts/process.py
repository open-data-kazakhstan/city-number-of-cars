import csv
from datapackage import Package
from openpyxl import load_workbook

def convert_xlsx_to_csv(input_file, output_file):
    wb = load_workbook(filename=input_file)
    ws = wb.active

    region_mapping = {
        'Республика Казахстан': 'The Republic Of Kazakhstan',
        'Абай': 'Abay Region',
        'Акмола': 'Akmola Region',
        'Актобе': 'Aktobe Region',
        'Алматы': 'Almaty Region',
        'Атырау': 'Atyrau Region',
        'Западный Казахстан': 'West Kazakhstan Region',
        'Джамбул': 'Zhambyl Region',
        'Жетысу': 'Zhetysu Region',
        'Караганда': 'Karaganda Region',
        'Костанай': 'Kostanay Region',
        'Кызылорда': 'Kyzylorda Region',
        'Мангистау': 'Mangystau Region',
        'Южно-Казахстан': 'South Kazakhstan Region',
        'Павлодар': 'Pavlodar Region',
        'Северо-Казахстан': 'North Kazakhstan Region',
        'Туркестан': 'Turkestan Region',
        'Улытау': 'Ulytau Region',
        'Восточно-Казахстанская': 'East Kazakhstan Region',
        'Г. Астана': 'Astana city',
        'Г. Алматы': 'Almaty city',
        'Г. Шымкент': 'Shymkent city'
    }

    with open(output_file, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)    
        for row in ws.iter_rows(values_only=True):
            region = row[0]
            value = row[1]
            if region in region_mapping:
                writer.writerow([region_mapping[region], value])

convert_xlsx_to_csv("archive/car2022.xlsx", "data/car2022.csv")

#datapackege

package = Package()
package.infer(r"data/car2022.csv")
package.commit()
package.save(r"datapackage.json")

